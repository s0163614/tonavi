from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
import os
import shutil
import json
import time
import logging
import openpyxl
from openpyxl.utils import get_column_letter
from functools import wraps
import uuid
import re
import requests
from datetime import datetime
import mysql.connector
from mysql.connector import Error
import math
from openpyxl.styles import numbers
from io import BytesIO
from orders import orders_bp  
from glob import glob
from urllib.parse import quote
import threading

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/images'
app.config['EXCEL_FILE'] = 'Каталог.xlsx'
app.secret_key = 'your-secret-key-here'

# MySQL configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_PORT'] = 3306
app.config['MYSQL_USER'] = 'tonavi_root'
app.config['MYSQL_PASSWORD'] = 'Ghjdjrfwbz2020'
app.config['MYSQL_DB'] = 'tonavi_root'
app.config['DEBUG'] = True
app.register_blueprint(orders_bp)


DADATA_API_KEY = "33ce2ae14246ae3bc798fb5495d8b1a04675e2e6"
DADATA_API_URL = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/findById/party"
# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)





def get_db_connection():
    try:
        conn = mysql.connector.connect(
            host=app.config['MYSQL_HOST'],
            port=app.config['MYSQL_PORT'],
            user=app.config['MYSQL_USER'],
            password=app.config['MYSQL_PASSWORD'],
            database=app.config['MYSQL_DB']
        )
        return conn
    except Error as e:
        print(f"Error connecting to MySQL: {e}")
        return None

users = {}


@app.template_filter('tojson')
def tojson_filter(obj):
    return json.dumps(obj, ensure_ascii=False)




def fetch_wb_data(article):
    API_KEY = "eyJhbGciOiJFUzI1NiIsImtpZCI6IjIwMjUwMjE3djEiLCJ0eXAiOiJKV1QifQ.eyJlbnQiOjEsImV4cCI6MTc1OTIxMTcwNywiaWQiOiIwMTk1ZWQ1Ny1jNmFjLTdjNzUtOGM5Ni1kYTZlMWEzNWU3NDgiLCJpaWQiOjQ0NDQ2MjE3LCJvaWQiOjg3NzA4LCJzIjo0Miwic2lkIjoiOWQxMWZmMDctYjZmZC01OGU1LThkMTAtYTlhYzIxMTIzZWExIiwidCI6ZmFsc2UsInVpZCI6NDQ0NDYyMTd9.kre5ByBTjNNVFDP7Y3DkKjx7GFJwIaanuW1OlRZab7bDf4lxZlDKm6VIxVyHbeTHxuPhdUAL1n4pjjyLVT4RFA"  # Замени на свой токен
    url = "https://content-api.wildberries.ru/content/v2/get/cards/list"

    headers = {
        "Authorization": API_KEY,
        "Content-Type": "application/json"
    }

    payload = {
        "settings": {
            "filter": {
                "textSearch": article,
                "withPhoto": 1
            },
            "cursor": {
                "limit": 1
            }
        }
    }

    response = requests.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        data = response.json()
        if data.get("cards"):
            card = data["cards"][0]
            description = card.get("description", "Описание отсутствует")
            weight_kg = card.get("dimensions", {}).get("weightBrutto", 0)
            weight_grams = int(weight_kg * 1000) if weight_kg else None
            return description, weight_grams
    return None, None


def get_local_product_images(article):
    """Ищем фото по артикулу ДО точки — чисто как 'МП95' -> МП95.jpg"""
    images_dir = os.path.join(os.path.dirname(__file__), 'static', 'foto')
    all_files = glob(os.path.join(images_dir, '*.*'))  # Все файлы с расширениями

    supported_formats = ('.jpg', '.jpeg', '.png', '.webp', '.JPG', '.JPEG', '.PNG', '.WEBP')

    for file_path in all_files:
        file_name = os.path.basename(file_path)
        name_without_ext = os.path.splitext(file_name)[0]
        if name_without_ext.strip().lower() == article.strip().lower():
            if file_path.lower().endswith(supported_formats):
                return f"foto/{file_name}"

    return None



def load_wb_commissions():
    categories = {}  # subjectName -> subjectID
    commissions = {}  # subjectName -> комиссия
    
    try:
        with open('commissions (2).txt', 'r', encoding='utf-8') as f:
            data = json.load(f)
            
            for item in data.get("report", []):
                # Очищаем название от невидимых символов и лишних пробелов
                name = item["subjectName"].strip()
                name = re.sub(r'[\u200b\u200e\u200f]', '', name)  # Удаляем невидимые символы
                name = ' '.join(name.split())  # Удаляем двойные пробелы
                
                categories[name] = item["subjectID"]
                commissions[name] = item.get("kgvpMarketplace", 0.0)
                
                # Дополнительно сохраняем вариант без кавычек
                name_unquoted = name.replace('"', '').replace("'", "")
                if name_unquoted != name:
                    categories[name_unquoted] = item["subjectID"]
                    commissions[name_unquoted] = item.get("kgvpMarketplace", 0.0)
                    
    except Exception as e:
        print(f"Ошибка загрузки комиссий: {str(e)}")
        # Для отладки выведем первые 10 категорий
        if 'data' in locals():
            print("Пример категорий в файле:", list(categories.keys())[:10])
    
    return categories, commissions


WB_CATEGORIES, WB_COMMISSIONS = load_wb_commissions()


@app.route('/wb_cards')
def wb_cards():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    file_path = os.path.join(app.root_path, app.config['EXCEL_FILE'])
    if not os.path.exists(file_path):
        return render_template('error.html', message="Файл каталога не найден"), 404

    try:
        products = parse_excel(file_path)
        return render_template('wb_cards.html',
                               products=products,
                               wb_categories=WB_CATEGORIES,
                               user={'id': session['user_id']})
    except Exception as e:
        logger.error(f"Ошибка загрузки каталога: {str(e)}")
        return render_template('error.html', message="Ошибка загрузки каталога"), 500









def upload_media_background(api_key, media_upload_queue):
    headers = {
        'Authorization': api_key,
        'Content-Type': 'application/json'
    }

    def get_all_cards():
        cursor = None
        all_cards = []
        while True:
            request_data = {
                "settings": {
                    "cursor": {
                        "limit": 100,
                        "cursor": cursor if cursor else {}
                    },
                    "filter": {
                        "withPhoto": -1
                    }
                }
            }
            try:
                cards_response = requests.post(
                    'https://content-api.wildberries.ru/content/v2/get/cards/list',
                    headers=headers,
                    json=request_data
                )
                cards_data = cards_response.json()
                all_cards.extend(cards_data.get('cards', []))
                cursor = cards_data.get('cursor', {}).get('updatedAt')
                if not cursor or len(cards_data.get('cards', [])) < 100:
                    break
            except Exception as e:
                print(f"❌ Ошибка при получении списка карточек: {str(e)}")
                break
        return all_cards

    all_cards = get_all_cards()

    for vendor_code, media_links in media_upload_queue:
        if not media_links:
            continue

        attempt = 0
        max_attempts = 5
        nmid = None

        while attempt < max_attempts:
            nmid = next((card.get('nmID') for card in all_cards if card.get('vendorCode', '').strip() == vendor_code), None)
            if nmid:
                break
            else:
                print(f"⚠️ Попытка {attempt+1}: nmID не найден для {vendor_code}. Повторная попытка через 2 секунды...")
                time.sleep(2)
                all_cards = get_all_cards()
                attempt += 1

        if not nmid:
            print(f"❌ Не найден nmID для товара с vendorCode {vendor_code} после {max_attempts} попыток.")
            continue

        upload_attempt = 0
        while upload_attempt < max_attempts:
            try:
                media_body = {
                    "nmId": nmid,
                    "data": media_links
                }
                media_response = requests.post(
                    'https://content-api.wildberries.ru/content/v3/media/save',
                    headers=headers,
                    json=media_body
                )
                if media_response.status_code == 200:
                    print(f"✅ Медиа добавлены в карточку {nmid}")
                    break
                else:
                    print(f"⚠️ Попытка {upload_attempt+1}: Ошибка загрузки медиа для {nmid}: {media_response.text}")
                    time.sleep(2)
                    upload_attempt += 1
            except Exception as e:
                print(f"❌ Ошибка при загрузке медиа для {nmid}: {str(e)}")
                time.sleep(2)
                upload_attempt += 1

        if upload_attempt == max_attempts:
            print(f"❌ Не удалось загрузить медиа для {nmid} после {max_attempts} попыток.")

@app.route('/create_wb_cards', methods=['POST'])
def create_wb_cards():
    try:
        selected_products = request.json.get('products', [])
        api_key = request.json.get('api_key', '')

        if not api_key:
            return jsonify({'success': False, 'message': 'API ключ не указан'}), 400

        # Загружаем JSON с данными WB
        json_path = os.path.join(os.path.dirname(__file__), 'static', 'wb_all_cards.json')
        with open(json_path, 'r', encoding='utf-8') as f:
            wb_data = json.load(f)

        # Создаем словари для быстрого поиска
        wb_dict = {item['vendorCode'].strip(): item for item in wb_data}
        media_upload_queue = []  # Для фоновой загрузки медиа
        cards_to_create = []
        errors = []

        for product in selected_products:
            article = product.get('article', '').strip()
            wb_item = wb_dict.get(article)

            if not wb_item:
                errors.append(f"Товар с артикулом '{article}' не найден в JSON")
                continue

            # Получаем subjectID из JSON или ищем по категории
            subject_id = wb_item.get('subjectID')
            if not subject_id:
                subject_name = wb_item.get('subjectName', '')
                # Нормализуем название категории для поиска
                normalized_name = re.sub(r'[^\w\s]', '', subject_name).strip().lower()
                for cat_name, cat_id in WB_CATEGORIES.items():
                    if normalized_name in cat_name.lower():
                        subject_id = cat_id
                        break
            
            if not subject_id:
                errors.append(f"Не найден subjectID для товара {article}")
                continue

            # Получаем фото для фоновой загрузки
            photos = [photo['big'] for photo in wb_item.get('photos', []) if 'big' in photo]
            if photos:
                media_upload_queue.append((article, photos))

            # Формируем варианты товара
            variants = []
            for variant in product.get('variants', []):
                try:
                    dimensions = wb_item.get('dimensions', {})
                    
                    # Рассчитываем цену с учетом комиссии
                    commission = WB_COMMISSIONS.get(wb_item.get('subjectName'), 16.5) / 100
                    base_price = float(variant.get('Цена', 0))
                    final_price = calculate_wb_price(
                        base_price=base_price,
                        dimensions=dimensions,
                        commission=commission,
                        tax=variant.get('Налог', 0.07),
                        limit=variant.get('Лимит', True),
                        margin=variant.get('Маржа', 0.20)  # Добавлен параметр маржи
                    )
                    
                    variants.append({
                        "vendorCode": variant.get('Артикул', article),
                        "title": wb_item.get('title', product.get('name', '')),
                        "description": wb_item.get('description', ''),
                        # Добавляем характеристики из JSON
                        "characteristics": wb_item.get('characteristics', []),
                        "dimensions": {
                            "length": dimensions.get('length', 10),
                            "width": dimensions.get('width', 10),
                            "height": dimensions.get('height', 10),
                            "weightBrutto": dimensions.get('weightBrutto', 0.3)
                        },
                        "sizes": [{
                            "price": final_price
                        }],
                        "mediaFiles": photos[:5]
                    })
                except Exception as e:
                    errors.append(f"Ошибка в варианте {variant.get('Артикул')}: {str(e)}")
                    continue

            if variants:
                cards_to_create.append({
                    "subjectID": subject_id,
                    "variants": variants
                })

        if not cards_to_create:
            return jsonify({
                'success': False,
                'message': 'Нет валидных товаров для создания',
                'errors': errors
            }), 400

        # Отправляем карточки в WB API
        headers = {'Authorization': api_key, 'Content-Type': 'application/json'}
        response = requests.post(
            'https://content-api.wildberries.ru/content/v2/cards/upload',
            headers=headers,
            json=cards_to_create
        )

        # Запускаем фоновую загрузку медиа
        if media_upload_queue:
            threading.Thread(
                target=upload_media_background,
                args=(api_key, media_upload_queue)
            ).start()

        if response.status_code == 200:
            return jsonify({
                'success': True,
                'message': 'Карточки успешно созданы! Медиа загружаются в фоне.',
                'response': response.json(),
                'warnings': errors if errors else None
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Ошибка API Wildberries',
                'status_code': response.status_code,
                'response_text': response.text,
                'errors': errors
            }), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Ошибка: {str(e)}',
            'traceback': traceback.format_exc()
        }), 500

def calculate_wb_price(base_price, dimensions, commission, tax, limit, margin):
    """Рассчитывает финальную цену для WB с учетом всех параметров"""
    AE2 = 60  # Логистика
    AC2 = 0.9  # Коэффициент
    K2 = 0.02 if limit else 0  # Процент 2
    AF2 = 0.015  # Процент 3
    AG2 = 0.015  # Процент 4
    L2 = margin  # Используем переданное значение маржи
    
    volume = (dimensions.get('length', 10) * 
             dimensions.get('width', 10) * 
             dimensions.get('height', 10)) / 1000
    
    ab2 = 43.75 + 10.625 * (math.ceil(volume) - 1)
    ad2 = (ab2 + (50 * (1 - AC2))) / AC2
    final_price = (base_price + AE2 + ad2) / (1 - tax - K2 - AF2 - AG2 - commission - L2)
    
    return round(final_price)



class User:
    def __init__(self, user_id, username, password, post=False, 
                 company_info=None, saved_companies=None, 
                 user_info=None, api_key=''):
        self.id = user_id
        self.username = username
        self.password = password
        self.post = post
        self.company_info = company_info or {}
        self.saved_companies = saved_companies or []
        self.user_info = user_info or ""
        self.api_key = api_key 

    def add_saved_company(self, company_data):
        self.saved_companies.append(company_data)
        self._update_db()

    def remove_saved_company(self, inn):
        self.saved_companies = [c for c in self.saved_companies if c.get('inn') != inn]
        self._update_db()

    def update_user_info(self, info):
        self.user_info = info
        self._update_db()

    def _update_db(self):
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                query = """
                UPDATE users 
                SET company_info = %s, saved_companies = %s, user_info = %s
                WHERE id = %s
                """
                cursor.execute(query, (
                    json.dumps(self.company_info) if self.company_info else None,
                    json.dumps(self.saved_companies) if self.saved_companies else None,
                    self.user_info,
                    self.id
                ))
                conn.commit()
            except Error as e:
                print(f"Error updating user in DB: {e}")
            finally:
                if conn.is_connected():
                    cursor.close()
                    conn.close()

# Обновленные маршруты регистрации и входа
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        post = request.form.get('post', '0') == '1'  # Получаем значение роли

        if not username or not password:
            return render_template('register.html', error='Заполните все поля')

        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor()
                
                cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
                if cursor.fetchone():
                    return render_template('register.html', error='Пользователь уже существует')

                user_id = str(uuid.uuid4())
                cursor.execute(
                    "INSERT INTO users (id, username, password, post) VALUES (%s, %s, %s, %s)",
                    (user_id, username, password, post)
                )
                conn.commit()

                user = User(user_id, username, password, post)
                session['user_id'] = user.id
                session['user_post'] = user.post  # Сохраняем роль в сессии
                return redirect(url_for('profile'))

            except Error as e:
                print(f"Error during registration: {e}")
                return render_template('register.html', error='Ошибка при регистрации')
            finally:
                if conn.is_connected():
                    cursor.close()
                    conn.close()
        else:
            return render_template('register.html', error='Ошибка подключения к базе данных')

    return render_template('register.html')


# Добавим новые маршруты



@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor(dictionary=True)
                cursor.execute(
                    "SELECT id, username, password, post, company_info, saved_companies FROM users WHERE username = %s",
                    (username,)
                )
                user_data = cursor.fetchone()

                if not user_data or user_data['password'] != password:
                    return render_template('login.html', error='Неверные учетные данные')

                user = User(
                    user_data['id'],
                    user_data['username'],
                    user_data['password'],
                    user_data['post'],
                    user_data.get('company_info'),
                    user_data.get('saved_companies')
                )
                session['user_id'] = user.id
                session['user_post'] = user.post  # Сохраняем роль в сессии
                return redirect(url_for('index'))

            except Error as e:
                print(f"Error during login: {e}")
                return render_template('login.html', error='Ошибка при входе')
            finally:
                if conn.is_connected():
                    cursor.close()
                    conn.close()
        else:
            return render_template('login.html', error='Ошибка подключения к базе данных')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('login'))

@app.route('/profile')
def profile():
    if 'user_id' not in session:
        print("User not in session, redirecting to login")
        return redirect(url_for('login'))

    try:
        conn = get_db_connection()
        if not conn:
            raise Exception("Database connection failed")
        
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            """SELECT id, username, post, company_info, 
               saved_companies, user_info, api_key 
               FROM users WHERE id = %s""",
            (session['user_id'],)
        )
        user_data = cursor.fetchone()

        if not user_data:
            print("User data not found")
            session.clear()
            return redirect(url_for('login'))

        # Обработка JSON данных с защитой от ошибок
        def safe_json_load(data):
            try:
                return json.loads(data) if isinstance(data, str) else data
            except json.JSONDecodeError:
                return {}

        company_info = safe_json_load(user_data.get('company_info'))
        saved_companies = safe_json_load(user_data.get('saved_companies')) or []

        user = User(
            user_data['id'],
            user_data['username'],
            '',
            bool(user_data.get('post', False)),
            company_info,
            saved_companies,
            user_data.get('user_info', ''),
            user_data.get('api_key', '')
        )
        
        return render_template('profile.html', user=user)

    except Exception as e:
        print(f"Error in profile route: {str(e)}")
        session.clear()
        return redirect(url_for('login'))
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()
        
@app.route('/save_api_key', methods=['POST'])
def save_api_key():
    if 'user_id' not in session:
        return jsonify({'success': False, 'message': 'Необходима авторизация'})

    api_key = request.form.get('api_key', '').strip()
    
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE users SET api_key = %s WHERE id = %s",
                (api_key, session['user_id'])
            )
            conn.commit()
            return jsonify({'success': True, 'message': 'API-ключ успешно сохранен'})
        except Error as e:
            print(f"Error saving API key: {e}")
            return jsonify({'success': False, 'message': 'Ошибка при сохранении API-ключа'})
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    else:
        return jsonify({'success': False, 'message': 'Ошибка подключения к базе данных'})
    
@app.route('/get_company_info', methods=['POST'])
def get_company_info():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Необходима авторизация"}), 401

    inn_data = request.get_json()
    if not inn_data:
        return jsonify({"success": False, "message": "Неверный формат запроса"}), 400

    inn = inn_data.get('inn', '').strip()

    if not inn or not re.match(r'^\d{10,12}$', inn):
        return jsonify({"success": False, "message": "Неверный ИНН"}), 400

    try:
        headers = {
            "Authorization": f"Token {DADATA_API_KEY}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        payload = {"query": inn, "count": 20}

        response = requests.post(DADATA_API_URL, headers=headers, json=payload, timeout=10)
        response.raise_for_status()
        data = response.json()

        suggestions = data.get("suggestions", [])
        if not suggestions:
            return jsonify({"success": False, "message": "Компания не найдена"})

        companies = []
        for item in suggestions:
            company = {
                "name": item.get("value", ""),
                "inn": item["data"].get("inn", ""),
                "ogrn": item["data"].get("ogrn", ""),
                "kpp": item["data"].get("kpp", ""),
                "status": item["data"].get("state", {}).get("status", ""),
                "registration_date": item["data"].get("state", {}).get("registration_date", ""),
                "address": item["data"].get("address", {}).get("value", ""),
                "director": item["data"].get("management", {}).get("name", ""),
                "director_position": item["data"].get("management", {}).get("post", ""),
                "okved": item["data"].get("okved", ""),
                "source": "Dadata"
            }
            companies.append(company)

        return jsonify({
            "success": True,
            "multiple": len(companies) > 1,
            "data": companies if len(companies) > 1 else companies[0]
        })
    except Exception as e:
        print(f"Ошибка при запросе к Dadata: {str(e)}")
        return jsonify({"success": False, "message": "Ошибка сервера при запросе к Dadata"})

@app.route('/save_company', methods=['POST'])
def save_company():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Необходима авторизация"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "message": "Ошибка подключения к базе данных"}), 500

    try:
        cursor = conn.cursor(dictionary=True)
        
        # Get current user
        cursor.execute(
            "SELECT saved_companies FROM users WHERE id = %s",
            (session['user_id'],)
        )
        user_data = cursor.fetchone()
        
        if not user_data:
            return jsonify({"success": False, "message": "Пользователь не найден"}), 404

        company_data = request.get_json()
        if not company_data:
            return jsonify({"success": False, "message": "Неверные данные"}), 400

        # Initialize saved_companies
        saved_companies = []
        if user_data['saved_companies']:
            try:
                saved_companies = json.loads(user_data['saved_companies']) if isinstance(user_data['saved_companies'], str) else user_data['saved_companies']
            except json.JSONDecodeError:
                saved_companies = []
        
        # Check if company already exists
        if any(c.get('inn') == company_data.get('inn') for c in saved_companies):
            return jsonify({"success": False, "message": "Компания уже сохранена"}), 400

        saved_companies.append(company_data)
        
        # Update database record
        cursor.execute(
            "UPDATE users SET saved_companies = %s WHERE id = %s",
            (json.dumps(saved_companies, ensure_ascii=False, default=str), session['user_id'])
        )
        conn.commit()
        
        
        return jsonify({"success": True, "message": "Компания сохранена"})

    except Exception as e:
        print(f"Error saving company: {e}")
        return jsonify({"success": False, "message": "Ошибка сервера"}), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/remove_company/<inn>', methods=['POST'])
def remove_company(inn):
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Необходима авторизация"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "message": "Ошибка подключения к базе данных"}), 500

    try:
        cursor = conn.cursor(dictionary=True)
        
        # Получаем текущего пользователя
        cursor.execute(
            "SELECT saved_companies FROM users WHERE id = %s",
            (session['user_id'],)
        )
        user_data = cursor.fetchone()
        
        if not user_data:
            return jsonify({"success": False, "message": "Пользователь не найден"}), 404

        saved_companies = user_data['saved_companies'] or []
        if isinstance(saved_companies, str):
            saved_companies = json.loads(saved_companies)
        
        # Удаляем компанию с указанным ИНН
        new_companies = [c for c in saved_companies if c.get('inn') != inn]
        
        # Обновляем запись в базе данных
        cursor.execute(
            "UPDATE users SET saved_companies = %s WHERE id = %s",
            (json.dumps(new_companies), session['user_id'])
        )
        conn.commit()
        
        return jsonify({"success": True, "message": "Компания удалена"})

    except Error as e:
        print(f"Error removing company: {e}")
        return jsonify({"success": False, "message": "Ошибка сервера"}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()



def cart_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'cart' not in session:
            session['cart'] = []  # Инициализируем корзину, если ее нет
        return f(*args, **kwargs)

    return decorated_function





@app.route('/add_to_cart', methods=['POST'])
@cart_required
def add_to_cart():
    data = request.get_json()
    row = data.get('row')
    
    file_path = os.path.join(app.root_path, app.config['EXCEL_FILE'])
    excel_data = get_excel_data(file_path)
    product = next((item for item in excel_data if item['row'] == row), None)
    
    if not product:
        return jsonify({'success': False, 'message': 'Товар не найден'})
    
    try:
        price = float(product['Цена'])
        if price <= 0:
            raise ValueError("Цена должна быть положительным числом")
    except (ValueError, TypeError) as e:
        logger.error(f"Ошибка преобразования цены: {str(e)}")
        return jsonify({'success': False, 'message': 'Неверный формат цены'})
    
    if not any(item['row'] == row for item in session['cart']):
        # Добавляем информацию о поставщике, если есть
        supplier_id = session.get('current_supplier_id')
        cart_item = {
            'row': product['row'],
            'Название': product['Название'],
            'Артикул': product['Артикул'],
            'Цена': price,
            'quantity': 1
        }
        if supplier_id:
            cart_item['supplier_id'] = supplier_id
            print(f"Added item from supplier: {supplier_id}") 
        
        session['cart'].append(cart_item)
        session.modified = True
        return jsonify({'success': True, 'cart_count': len(session['cart'])})
    
    return jsonify({'success': False, 'message': 'Товар уже в корзине'})

def get_supplier_name(supplier_id):
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT username FROM users WHERE id = %s", (supplier_id,))
            supplier = cursor.fetchone()
            return supplier['username'] if supplier else "Неизвестный поставщик"
        except Error as e:
            print(f"Error fetching supplier: {e}")
            return "Ошибка загрузки"
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
    return "Нет подключения"

@app.context_processor
def utility_processor():
    return dict(get_supplier_name=get_supplier_name)

@app.route('/confirm_order', methods=['POST'])
def confirm_order():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Необходима авторизация"}), 401
    
    # Only sellers (post=0) can confirm orders
    if session.get('user_post', True):
        return jsonify({"success": False, "message": "Только продавцы могут оформлять заказы"}), 403
    
    if 'cart' not in session or not session['cart']:
        return jsonify({"success": False, "message": "Корзина пуста"}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "message": "Ошибка подключения к БД"}), 500
    
    try:
        # Verify all items have the same supplier
        supplier_ids = {item.get('supplier_id') for item in session['cart']}
        if len(supplier_ids) != 1:
            return jsonify({"success": False, "message": "Все товары должны быть от одного поставщика"}), 400
        
        supplier_id = supplier_ids.pop()
        
        # Verify supplier exists
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id FROM users WHERE id = %s AND post = 1", (supplier_id,))
        if not cursor.fetchone():
            return jsonify({"success": False, "message": "Поставщик не найден"}), 404
        
        # Prepare order items and calculate total
        items = []
        total = 0.0
        for item in session['cart']:
            item_total = float(item['Цена']) * item['quantity']
            items.append({
                'name': item['Название'],
                'article': item['Артикул'],
                'price': float(item['Цена']),
                'quantity': item['quantity'],
                'row': item['row']
            })
            total += item_total
        
        # Convert to JSON string with ensure_ascii=False
        items_json = json.dumps(items, ensure_ascii=False)
        
        # Debug print
        print(f"Items JSON to be stored: {items_json}")
        print(f"Calculated total: {total}")
        
        # Create order
        order_id = str(uuid.uuid4())
        cursor.execute(
            "INSERT INTO orders (id, seller_id, supplier_id, items, total, status) VALUES (%s, %s, %s, %s, %s, 'pending')",
            (order_id, session['user_id'], supplier_id, items_json, total)
        )
        conn.commit()
        
        # Verify order was created
        cursor.execute("SELECT id FROM orders WHERE id = %s", (order_id,))
        if not cursor.fetchone():
            raise Exception("Order not created")
        
        # Clear cart
        session['cart'] = []
        session.modified = True
        
        return jsonify({
            "success": True, 
            "message": "Заказ успешно оформлен",
            "order_id": order_id
        })
    
    except Exception as e:
        conn.rollback()
        print(f"Error creating order: {e}")
        return jsonify({"success": False, "message": f"Ошибка при оформлении заказа: {str(e)}"}), 500
    finally:
        if conn.is_connected():
            conn.close()

@app.route('/confirm_payment/<order_id>', methods=['POST'])
def confirm_payment(order_id):
    if 'user_id' not in session or not session.get('user_post', False):
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        return "Ошибка подключения к базе данных", 500
    
    try:
        cursor = conn.cursor()
        # Verify the order belongs to this supplier
        cursor.execute(
            "UPDATE orders SET status = 'confirmed', confirmed_at = CURRENT_TIMESTAMP WHERE id = %s AND supplier_id = %s AND status = 'pending'",
            (order_id, session['user_id'])
        )
        conn.commit()
        
        if cursor.rowcount == 0:
            return jsonify({"success": False, "message": "Заказ не найден или уже обработан"}), 400
        
        return jsonify({"success": True, "message": "Платеж подтвержден"})
    
    except Error as e:
        print(f"Error confirming payment: {e}")
        return jsonify({"success": False, "message": "Ошибка сервера"}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/supplier/orders')
def supplier_orders():
    if 'user_id' not in session or session.get('user_post', 0) != 1:  # Только для поставщиков (post=1)
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        return render_template('error.html', message="Ошибка подключения к БД"), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Получаем информацию о продавцах для текущего поставщика
        cursor.execute("""
            SELECT DISTINCT u.id, u.username, COALESCE(SUM(d.amount), 0) as deposit_amount
            FROM users u
            JOIN deposits d ON u.id = d.seller_id
            WHERE d.supplier_id = %s AND u.post = 0
            GROUP BY u.id, u.username
        """, (session['user_id'],))
        
        sellers = {seller['id']: seller for seller in cursor.fetchall()}
        
        cursor.execute("""
            SELECT 
                o.id,
                o.status,
                o.seller_id,
                o.supplier_id,
                o.total,
                o.items AS order_items,
                o.created_at,
                DATE_FORMAT(o.created_at, '%%d.%%m.%%Y %%H:%%i') as created_at_formatted,
                o.confirmed_at,
                u.username as seller_name,
                DATE_FORMAT(o.confirmed_at, '%%d.%%m.%%Y %%H:%%i') as confirmed_at_formatted,
                COALESCE(d.deposit_sum, 0) AS seller_deposit
            FROM orders o
            JOIN users u ON o.seller_id = u.id
            LEFT JOIN (
                SELECT seller_id, SUM(amount) AS deposit_sum
                FROM deposits
                WHERE supplier_id = %s
                GROUP BY seller_id
            ) d ON o.seller_id = d.seller_id
            WHERE o.supplier_id = %s
            ORDER BY o.created_at DESC
        """, (session['user_id'], session['user_id']))
        
        orders = []
        for row in cursor.fetchall():
            order = dict(row)
            items_data = order.pop('order_items', '[]')
            try:
                order['order_items'] = json.loads(items_data) if isinstance(items_data, str) else items_data
            except json.JSONDecodeError:
                order['order_items'] = []
            orders.append(order)

        return render_template('supplier_orders.html', orders=orders, sellers=sellers)
    
    except Error as e:
        print(f"Error fetching orders: {e}")
        return render_template('error.html', message="Ошибка загрузки заказов"), 500
    finally:
        if conn.is_connected():
            conn.close()

@app.route('/supplier/deposit/<seller_id>', methods=['GET', 'POST'])
def reduce_deposit(seller_id):
    # Проверка авторизации и прав доступа (только для поставщиков)
    if 'user_id' not in session or session.get('user_post', 0) != 1:
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        return render_template('error.html', message="Ошибка подключения к БД"), 500

    try:
        cursor = conn.cursor(dictionary=True)

        # Получаем данные о продавце и общем депозите
        cursor.execute("""
            SELECT 
                u.id, 
                u.username, 
                COALESCE(SUM(d.amount), 0) as total_deposit
            FROM users u
            LEFT JOIN deposits d ON u.id = d.seller_id AND d.supplier_id = %s
            WHERE u.id = %s AND u.post = 0
            GROUP BY u.id
        """, (session['user_id'], seller_id))
        
        seller = cursor.fetchone()

        if not seller:
            return render_template('error.html', message="Продавец не найден или не является продавцом"), 404

        if request.method == 'POST':
            reduce_amount = request.form.get('reduce_amount', '').strip()
            
            # Валидация введенной суммы
            if not reduce_amount:
                return render_template('error.html', message="Введите сумму для уменьшения"), 400
                
            try:
                reduce_amount = float(reduce_amount)
                if reduce_amount <= 0:
                    return render_template('error.html', message="Сумма должна быть больше нуля"), 400
                if reduce_amount > float(seller['total_deposit']):
                    return render_template('error.html', message="Нельзя уменьшить больше, чем текущий депозит"), 400
                if reduce_amount > 1000000:  # Пример ограничения максимальной суммы
                    return render_template('error.html', message="Сумма слишком большая"), 400
            except ValueError:
                return render_template('error.html', message="Введите корректную сумму"), 400

            # Добавляем запись об уменьшении депозита (с отрицательным значением)
            cursor.execute("""
                INSERT INTO deposits (seller_id, supplier_id, amount) 
                VALUES (%s, %s, %s)
            """, (seller_id, session['user_id'], -reduce_amount))

            conn.commit()

            # Возвращаем сообщение об успехе
            return render_template(
                'success.html',
                message=f"Депозит от продавца {seller['username']} уменьшен на {reduce_amount:.2f} ₽"
            )

        # GET запрос - отображаем форму
        return render_template(
            'reduce_deposit.html',
            seller={
                'id': seller['id'],
                'username': seller['username'],
                'total_deposit': float(seller['total_deposit'])  # Гарантируем float тип
            }
        )

    except Exception as e:
        print(f"Error in reduce_deposit: {str(e)}")
        return render_template('error.html', message="Произошла ошибка при обработке запроса"), 500
    finally:
        if conn and conn.is_connected():
            conn.close()

@app.route('/checkout', methods=['GET'])
def checkout():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('user_post', True):  # Only sellers (post=0) can checkout
        return redirect(url_for('index'))
    
    if 'cart' not in session or not session['cart']:
        return redirect(url_for('view_cart'))
    
    conn = get_db_connection()
    if not conn:
        return "Ошибка подключения к базе данных", 500
    
    try:
        # Get seller info (current user)
        cursor = conn.cursor(dictionary=True)
        cursor.execute(
            "SELECT id, username, user_info FROM users WHERE id = %s",
            (session['user_id'],)
        )
        seller = cursor.fetchone()
        
        # Get supplier info from first item in cart
        supplier_id = session['cart'][0].get('supplier_id')
        if not supplier_id:
            return "Не удалось определить поставщика", 400
            
        cursor.execute(
            "SELECT id, username, user_info FROM users WHERE id = %s",
            (supplier_id,)
        )
        supplier = cursor.fetchone()
        
        if not seller or not supplier:
            return "Не удалось получить информацию о пользователях", 404
        
        return render_template('checkout.html', 
                             seller=seller,
                             supplier=supplier,
                             cart_items=session['cart'],
                             total=sum(item['Цена'] * item['quantity'] for item in session['cart']))
        
    except Error as e:
        print(f"Error during checkout: {e}")
        return "Ошибка сервера", 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/cart')
@cart_required
def view_cart():
    cart_items = session.get('cart', [])
    # Убедимся, что все цены - числа
    for item in cart_items:
        if isinstance(item['Цена'], str):
            try:
                item['Цена'] = float(item['Цена'])
            except (ValueError, TypeError):
                item['Цена'] = 0.0
    
    total = sum(item['Цена'] * item['quantity'] for item in cart_items)
    
    # Получаем информацию о пользователе
    user = None
    if 'user_id' in session:
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor(dictionary=True)
                cursor.execute("SELECT id, username, post FROM users WHERE id = %s", (session['user_id'],))
                user = cursor.fetchone()
            except Exception as e:
                logger.error(f"Ошибка получения данных пользователя: {str(e)}")
            finally:
                if conn and conn.is_connected():
                    cursor.close()
                    conn.close()
    
    return render_template('cart.html', 
                         cart_items=cart_items, 
                         total=f"{total:.2f} ₽",
                         user=user)  # Передаем данные пользователя в шаблон

@app.route('/remove_from_cart/<int:row>', methods=['POST'])
@cart_required
def remove_from_cart(row):
    session['cart'] = [item for item in session['cart'] if item['row'] != row]
    session.modified = True
    return redirect(url_for('view_cart'))

@app.route('/supplier_products/<int:supplier_id>')
def supplier_products(supplier_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Получаем данные текущего пользователя
    conn = get_db_connection()
    if not conn:
        return render_template('error.html', message="Ошибка подключения к БД"), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Проверяем, существует ли поставщик
        cursor.execute("SELECT id, username, post FROM users WHERE id = %s AND post = 1", (supplier_id,))
        supplier = cursor.fetchone()
        
        if not supplier:
            return render_template('error.html', message="Поставщик не найден"), 404
        
        # Получаем данные текущего пользователя
        cursor.execute("SELECT id, username, post FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()
        
        if not user:
            session.pop('user_id', None)
            return redirect(url_for('login'))
        
        # Сохраняем ID поставщика в сессии
        session['current_supplier_id'] = supplier_id
        
        # Получаем параметры фильтрации
        search_query = request.args.get('query', '').strip()
        selected_categories = request.args.getlist('category')
        
        # Загружаем товары поставщика
        file_path = os.path.join(app.root_path, app.config['EXCEL_FILE'])
        if not os.path.exists(file_path):
            return render_template('error.html', message="Файл каталога не найден"), 404
        
        try:
            products = parse_excel(file_path)
            
            # Фильтрация по поисковому запросу
            if search_query:
                search_lower = search_query.lower()
                products = [p for p in products if search_lower in p['name'].lower()]
            
            # Фильтрация по выбранным категориям
            if selected_categories:
                products = [p for p in products if p.get('category') in selected_categories]
            
            # Получаем список всех уникальных категорий для фильтра
            all_categories = sorted({p.get('category') for p in products if p.get('category')})
            
            return render_template('index.html', 
                                products=products,
                                user=user,
                                is_seller=not user['post'],  # True если продавец
                                supplier_view=True,
                                supplier=supplier,
                                all_categories=all_categories,
                                selected_categories=selected_categories)
        except Exception as e:
            logger.error(f"Ошибка загрузки каталога: {str(e)}")
            return render_template('error.html', message="Ошибка загрузки каталога"), 500
            
    except Exception as e:
        logger.error(f"Ошибка в supplier_products: {str(e)}")
        return render_template('error.html', message="Внутренняя ошибка сервера"), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()


@app.route('/seller/orders')
def seller_orders():
    if 'user_id' not in session or session.get('user_post', 0) != 0:
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    if not conn:
        return render_template('error.html', message="Ошибка подключения к БД"), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Получаем информацию о поставщиках для текущего продавца
        cursor.execute("""
            SELECT DISTINCT u.id, u.username, COALESCE(u.sent_amount, 0) as sent_amount
            FROM users u
            JOIN orders o ON u.id = o.supplier_id
            WHERE o.seller_id = %s AND u.post = 1
        """, (session['user_id'],))
        
        suppliers = {supplier['id']: supplier for supplier in cursor.fetchall()}
        
        # Извлекаем заказы
        cursor.execute("""
            SELECT 
                o.id,
                o.status,
                o.seller_id,
                o.supplier_id,
                o.total,
                o.items AS order_items,
                o.created_at,
                DATE_FORMAT(o.created_at, '%%d.%%m.%%Y %%H:%%i') as created_at_formatted,
                o.confirmed_at,
                u.username as supplier_name,
                DATE_FORMAT(o.confirmed_at, '%%d.%%m.%%Y %%H:%%i') as confirmed_at_formatted,
                COALESCE(d.total_amount, 0) AS supplier_deposit
            FROM orders o
            JOIN users u ON o.supplier_id = u.id
            LEFT JOIN (
                SELECT supplier_id, SUM(amount) AS total_amount
                FROM deposits
                WHERE seller_id = %s
                GROUP BY supplier_id
            ) d ON o.supplier_id = d.supplier_id
            WHERE o.seller_id = %s
            ORDER BY o.created_at DESC
        """, (session['user_id'], session['user_id']))
        
        orders = []
        for row in cursor.fetchall():
            order = dict(row)
            items_data = order.pop('order_items', '[]')
            try:
                order['order_items'] = json.loads(items_data) if isinstance(items_data, str) else items_data
            except json.JSONDecodeError:
                order['order_items'] = []
            orders.append(order)

        return render_template('seller_orders.html', orders=orders, suppliers=suppliers)
    
    except Error as e:
        print(f"Error fetching orders: {e}")
        return render_template('error.html', message="Ошибка загрузки заказов"), 500
    finally:
        if conn.is_connected():
            conn.close()


@app.route('/seller/deposit/<supplier_id>', methods=['GET', 'POST'])
def update_deposit(supplier_id):
    # Проверка авторизации и прав доступа (только для продавцов)
    if 'user_id' not in session or session.get('user_post', 0) != 0:
        return redirect(url_for('login'))

    # Подключение к базе данных
    conn = get_db_connection()
    if not conn:
        return render_template('error.html', message="Ошибка подключения к БД"), 500

    try:
        cursor = conn.cursor(dictionary=True)

        # Получаем данные о поставщике с суммой депозита от текущего продавца
        cursor.execute("""
            SELECT 
                u.id, 
                u.username, 
                COALESCE((
                    SELECT SUM(amount) 
                    FROM deposits 
                    WHERE seller_id = %s AND supplier_id = u.id
                ), 0) AS deposit_amount
            FROM users u 
            WHERE u.id = %s AND u.post = 1
        """, (session['user_id'], supplier_id))
        
        supplier = cursor.fetchone()

        if not supplier:
            return render_template('error.html', message="Поставщик не найден или не является поставщиком"), 404

        if request.method == 'POST':
            deposit_amount = request.form.get('deposit_amount', '').strip()
            
            # Валидация введенной суммы
            if not deposit_amount:
                return render_template('error.html', message="Введите сумму депозита"), 400
                
            try:
                deposit_amount = float(deposit_amount)
                if deposit_amount <= 0:
                    return render_template('error.html', message="Сумма должна быть больше нуля"), 400
                if deposit_amount > 1000000:  # Пример ограничения максимальной суммы
                    return render_template('error.html', message="Сумма слишком большая"), 400
            except ValueError:
                return render_template('error.html', message="Введите корректную сумму"), 400

            # Добавление записи о депозите
            cursor.execute("""
                INSERT INTO deposits (seller_id, supplier_id, amount) 
                VALUES (%s, %s, %s)
            """, (session['user_id'], supplier_id, deposit_amount))

            conn.commit()

            # Возвращаем сообщение об успехе
            return render_template(
                'success.html',
                message=f"Депозит для поставщика {supplier['username']} пополнен на {deposit_amount:.2f} ₽"
            )

        # GET запрос - отображаем форму
        return render_template(
            'update_deposit.html',
            supplier={
                'id': supplier['id'],
                'username': supplier['username'],
                'deposit_amount': float(supplier['deposit_amount'])  # Используем deposit_amount вместо sent_amount
            }
        )

    except Exception as e:
        print(f"Error in update_deposit: {str(e)}")
        return render_template('error.html', message="Произошла ошибка при обработке запроса"), 500
    finally:
        if conn and conn.is_connected():
            conn.close()

@app.route('/update_cart_item/<int:row>', methods=['POST'])
@cart_required
def update_cart_item(row):
    quantity = int(request.form.get('quantity', 1))
    
    for item in session['cart']:
        if item['row'] == row:
            item['quantity'] = quantity
            break
    
    session.modified = True
    return redirect(url_for('view_cart'))




@app.route('/update_user_info', methods=['POST'])
def update_user_info():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Необходима авторизация"}), 401

    conn = get_db_connection()
    if not conn:
        return jsonify({"success": False, "message": "Ошибка подключения к базе данных"}), 500

    try:
        info = request.form.get('user_info', '')
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET user_info = %s WHERE id = %s",
            (info, session['user_id'])
        )
        conn.commit()
        return jsonify({"success": True, "message": "Информация обновлена"})
    except Error as e:
        print(f"Error updating user info: {e}")
        return jsonify({"success": False, "message": "Ошибка сервера"}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/clear_cart', methods=['POST'])
@cart_required
def clear_cart():
    session['cart'] = []
    session.modified = True
    return redirect(url_for('view_cart'))



@app.route('/chats')
def chats():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    is_supplier = session.get('user_post', 0)
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        
        if is_supplier:
            # Для поставщика: чаты где он участник
            cursor.execute('''
                SELECT c.id, u.id as partner_id, u.username as partner_name 
                FROM chats c
                JOIN users u ON c.seller_id = u.id
                WHERE c.supplier_id = %s
            ''', (user_id,))
        else:
            # Для продавца: список всех поставщиков
            cursor.execute('SELECT id, username FROM users WHERE post = 1')
            
        chats_list = cursor.fetchall()
        return render_template('chats.html', 
                             chats=chats_list, 
                             is_supplier=is_supplier)
    
    except Error as e:
        print(f"Error: {e}")
        return "Ошибка базы данных", 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/chat/<string:partner_id>')
def chat(partner_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Приведение типов к строке
    user_id = str(session['user_id'])
    partner_id = str(partner_id)
    is_supplier = session.get('user_post', 0)
    
    conn = get_db_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Логирование параметров
        app.logger.debug(f"Поиск чата: user_id={user_id}, partner_id={partner_id}")
        
        # Исправленный запрос
        cursor.execute('''
            SELECT id FROM chats 
            WHERE (seller_id = %s AND supplier_id = %s)
               OR (supplier_id = %s AND seller_id = %s)
        ''', (user_id, partner_id, user_id, partner_id))
        
        chat_data = cursor.fetchone()
        
        if not chat_data:
            if is_supplier:
                app.logger.warning("Чат не найден для поставщика")
                return "Чат не найден", 404
            
            # Создание нового чата
            new_chat_id = str(uuid.uuid4())
            cursor.execute('''
                INSERT INTO chats (id, seller_id, supplier_id)
                VALUES (%s, %s, %s)
            ''', (new_chat_id, user_id, partner_id))
            conn.commit()
            chat_id = new_chat_id
        else:
            chat_id = chat_data['id']
        
        # Получение сообщений
        cursor.execute('''
            SELECT m.*, u.username as sender_name
            FROM messages m
            JOIN users u ON m.sender_id = u.id
            WHERE chat_id = %s
            ORDER BY sent_at ASC
        ''', (chat_id,))
        messages = cursor.fetchall()
        
        return render_template('chat.html', 
                            messages=messages,
                            partner_id=partner_id,
                            chat_id=chat_id)
    
    except Error as e:
        app.logger.error(f"Ошибка БД: {str(e)}")
        return "Ошибка базы данных", 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/send_message', methods=['POST'])
def send_message():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        app.logger.debug(f"Received message data: {data}")
        
        chat_id = data.get('chat_id')
        content = data.get('content', '').strip()
        
        if not content:
            return jsonify({'status': 'error', 'message': 'Empty message'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO messages (id, chat_id, sender_id, content)
            VALUES (UUID(), %s, %s, %s)
        ''', (chat_id, session['user_id'], content))
        conn.commit()
        
        app.logger.info(f"Message saved: chat_id={chat_id}, sender_id={session['user_id']}")
        return jsonify({'status': 'success'})
    
    except Error as e:
        app.logger.error(f"Database error: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Database error'}), 500
    except Exception as e:
        app.logger.error(f"Unexpected error: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Internal error'}), 500
    finally:
        if 'conn' in locals() and conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/chat_messages')
def chat_messages():
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    chat_id = request.args.get('chat_id')
    
    conn = get_db_connection()
    try:
        # Проверка принадлежности чата
        cursor = conn.cursor(dictionary=True)
        cursor.execute('''
            SELECT * FROM chats 
            WHERE id = %s 
              AND (seller_id = %s OR supplier_id = %s)
        ''', (chat_id, session['user_id'], session['user_id']))
        
        if not cursor.fetchone():
            return jsonify({'error': 'Access denied'}), 403

        # Получение сообщений
        cursor.execute('''
            SELECT m.*, u.username as sender_name
            FROM messages m
            JOIN users u ON m.sender_id = u.id
            WHERE chat_id = %s
            ORDER BY sent_at ASC
        ''', (chat_id,))
        messages = cursor.fetchall()
        return jsonify([dict(msg) for msg in messages])
    
    except Error as e:
        print(f"Error: {e}")
        return jsonify({'error': 'Database error'}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()



def safe_excel_operation(file_path):
    """Безопасная обертка для работы с Excel"""
    pythoncom.CoInitialize()
    excel = None
    wb = None
    
    try:
        excel = win32com.client.DispatchEx("Excel.Application")  # Используем DispatchEx
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        
        # Полный путь к файлу
        abs_path = os.path.abspath(file_path)
        logger.info(f"Открываем файл: {abs_path}")
        
        wb = excel.Workbooks.Open(abs_path)
        sheet = wb.Sheets(1)
        
        yield sheet
        
    except Exception as e:
        logger.error(f"Ошибка в Excel: {e}")
        raise
    finally:
        try:
            if wb:
                wb.Close(False)
            if excel:
                excel.Quit()
        except Exception as e:
            logger.error(f"Ошибка при закрытии Excel: {e}")
        finally:
            pythoncom.CoUninitialize()






@app.route('/export_wb_template', methods=['POST'])
def export_wb_template():
    try:
        selected_products = request.json.get('products', [])

        template_path = os.path.join(app.root_path, 'Шаблон для ВБ.xlsx')
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        percent_format = numbers.FORMAT_PERCENTAGE_00
        row_num = 2

        for product in selected_products:
            for variant in product['variants']:
                ws[f'A{row_num}'] = variant.get('Артикул', '')
                ws[f'G{row_num}'] = product.get('name', '')
                ws[f'W{row_num}'] = variant.get('Длина (см)', '')
                ws[f'X{row_num}'] = variant.get('Ширина (см)', '')
                ws[f'Y{row_num}'] = variant.get('Высота (см)', '')
                ws[f'AA{row_num}'] = 300
                ws[f'H{row_num}'] = product.get('category', '')
                commission = WB_COMMISSIONS.get(product.get('category', ''), 16.5)
                ws[f'I{row_num}'].value = commission / 100
                ws[f'I{row_num}'].number_format = percent_format
                ws[f'S{row_num}'] = variant.get('Цена', '')
                row_num += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"WB_export_{timestamp}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500








@app.route('/create_from_template', methods=['POST'])
def create_from_template():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'Файл не загружен'}), 400

        file = request.files['file']
        api_key = request.form.get('api_key', '')

        if not api_key:
            return jsonify({'success': False, 'error': 'API ключ не указан'}), 400

        if not file.filename.endswith('.xlsx'):
            return jsonify({'success': False, 'error': 'Неверный формат файла'}), 400

        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        products = []
        current_product = None

        for row in range(2, ws.max_row + 1):
            article = ws[f'A{row}'].value
            name = ws[f'G{row}'].value
            category = ws[f'H{row}'].value
            price = ws[f'Q{row}'].value
            length = ws[f'W{row}'].value
            width = ws[f'X{row}'].value
            height = ws[f'Y{row}'].value
            weight = ws[f'AA{row}'].value

            if not article:
                continue

            try:
                price = float(price) if price else 0
                length = int(length) if length else 10
                width = int(width) if width else 10
                height = int(height) if height else 10
                weight = float(weight) if weight else 300
            except Exception as e:
                print(f"Ошибка преобразования в строке {row}: {e}")
                continue

            variant = {
                'Артикул': str(article),
                'Название': name,
                'Длина (см)': length,
                'Ширина (см)': width,
                'Высота (см)': height,
                'Вес (г)': weight,
                'Цена': price,
                'Категория': category
            }

            if current_product and current_product['article'] == article:
                current_product['variants'].append(variant)
            else:
                if current_product:
                    products.append(current_product)
                current_product = {
                    'article': article,
                    'name': name,
                    'category': category,
                    'variants': [variant]
                }

        if current_product:
            products.append(current_product)

        return create_wb_cards_from_uploaded_template(products, api_key)

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500




















def create_wb_cards_from_uploaded_template(products, api_key):
    try:
        cards_to_create = []

        for product in products:
            subject_id = WB_CATEGORIES.get(product['category'], 0)
            if not subject_id:
                print(f"Категория не найдена: {product['category']}")
                continue

            variants = []
            for variant in product['variants']:
                try:
                    variants.append({
                        "vendorCode": variant['Артикул'],
                        "title": product['name'],
                        "dimensions": {
                            "length": variant['Длина (см)'],
                            "width": variant['Ширина (см)'],
                            "height": variant['Высота (см)'],
                            "weightBrutto": variant['Вес (г)'] / 1000
                        },
                        "sizes": [{
                            "price": round(float(variant['Цена']))
                        }]
                    })
                except Exception as e:
                    print(f"Ошибка при создании варианта: {str(e)}")

            if variants:
                cards_to_create.append({
                    "subjectID": subject_id,
                    "variants": variants
                })

        headers = {
            'Authorization': api_key,
            'Content-Type': 'application/json'
        }

        response = requests.post(
            'https://content-api.wildberries.ru/content/v2/cards/upload',
            headers=headers,
            json=cards_to_create
        )

        if response.status_code == 200:
            return jsonify({
                'success': True,
                'message': 'Карточки отправлены на создание',
                'response': response.json()
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Ошибка при создании карточек',
                'status_code': response.status_code,
                'response': response.text
            }), 400

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Ошибка: {str(e)}'
        }), 500


def get_excel_data(file_path):
    """Получаем данные из Excel файла"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    
    for row in range(4, sheet.max_row + 1):
        try:
            price = sheet[f'H{row}'].value
            # Преобразуем цену в число, если возможно
            try:
                price = float(price) if price is not None else 0.0
            except (ValueError, TypeError):
                price = 0.0
                logger.warning(f"Некорректная цена в строке {row}, установлено 0")
            
            data.append({
                'row': row,
                'Название': sheet[f'C{row}'].value,
                'Артикул': sheet[f'D{row}'].value,
                'Длина (см)': sheet[f'E{row}'].value,
                'Ширина (см)': sheet[f'F{row}'].value,
                'Высота (см)': sheet[f'G{row}'].value,
                'Цена': price,  # Уже число
            })
        except Exception as e:
            logger.error(f"Ошибка в строке {row}: {str(e)}")
            continue
    
    wb.close()
    return data

def update_excel_row(file_path, row_data):
    """Обновляем строку в Excel"""
    try:
        # Проверяем доступность файла для записи
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл {file_path} не найден")
            
        if not os.access(file_path, os.W_OK):
            raise PermissionError(f"Нет прав на запись в файл {file_path}")

        # Создаем резервную копию
        backup_path = file_path + '.bak'
        shutil.copy2(file_path, backup_path)
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        row = row_data['row']
        
        sheet[f'C{row}'] = row_data['Название']
        sheet[f'D{row}'] = row_data['Артикул']
        sheet[f'E{row}'] = row_data['Длина (см)']
        sheet[f'F{row}'] = row_data['Ширина (см)']
        sheet[f'G{row}'] = row_data['Высота (см)']
        sheet[f'H{row}'] = row_data['Цена']
        
        # Сохраняем с временным именем, затем переименовываем
        temp_path = file_path + '.tmp'
        wb.save(temp_path)
        wb.close()
        
        # Атомарная замена файла
        os.replace(temp_path, file_path)
        
    except Exception as e:
        # Восстанавливаем из резервной копии при ошибке
        if os.path.exists(backup_path):
            os.replace(backup_path, file_path)
        raise
    finally:
        # Удаляем временные файлы
        if os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except:
                pass
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass

def delete_excel_row(file_path, row_num):
    """Удаляем строку из Excel"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    sheet.delete_rows(row_num)
    wb.save(file_path)
    wb.close()

def parse_excel(file_path):
    products = []
    current_group = None
    group_data = []
    article_set = set()

    # Загружаем JSON-файл с данными товаров
    try:
        json_path = os.path.join(os.path.dirname(__file__), 'static', 'wb_all_cards.json')
        with open(json_path, 'r', encoding='utf-8') as f:
            wb_data = json.load(f)
    except Exception as e:
        print(f"Ошибка при загрузке JSON: {str(e)}")
        return []

    # Создаём словарь товаров по vendorCode для быстрого поиска
    wb_dict = {item['vendorCode'].strip(): item for item in wb_data}

    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Собираем все артикулы для проверки
        for row in range(4, sheet.max_row + 1):
            article = sheet[f'D{row}'].value
            if article:
                article_set.add(str(article).strip())

        # Основной цикл обработки данных
        for row in range(4, sheet.max_row + 1):
            article = str(sheet[f'D{row}'].value).strip() if sheet[f'D{row}'].value else None
            if not article:
                continue

            # Обработка смены группы товаров
            if article != current_group:
                if group_data:
                    wb_item = wb_dict.get(current_group)
                    if wb_item:
                        # Извлекаем категорию из JSON
                        category = wb_item.get('subjectName', 'Категория не указана')
                        description = wb_item.get('description', 'Описание отсутствует')
                        photos = wb_item.get('photos', [])
                        big_photo = photos[0]['big'] if photos else None
                        weight = wb_item.get('dimensions', {}).get('weightBrutto', None)

                        products.append({
                            'article': current_group,
                            'name': group_data[0]['Название'],
                            'category': category,  # Используем категорию из JSON
                            'variants': group_data,
                            'image': big_photo,
                            'description': description,
                            'weight': weight,
                        })

                current_group = article
                group_data = []

            # Добавляем вариант товара
            if sheet[f'C{row}'].value:
                group_data.append({
                    'row': row,
                    'Название': sheet[f'C{row}'].value,
                    'Артикул': article,
                    'Длина (см)': sheet[f'E{row}'].value,
                    'Ширина (см)': sheet[f'F{row}'].value,
                    'Высота (см)': sheet[f'G{row}'].value,
                    'Цена': sheet[f'H{row}'].value,
                })

        # Обработка последней группы
        if group_data:
            wb_item = wb_dict.get(current_group)
            if wb_item:
                category = wb_item.get('subjectName', 'Категория не указана')
                description = wb_item.get('description', 'Описание отсутствует')
                photos = wb_item.get('photos', [])
                big_photo = photos[0]['big'] if photos else None
                weight = wb_item.get('dimensions', {}).get('weightBrutto', None)

                products.append({
                    'article': current_group,
                    'name': group_data[0]['Название'],
                    'category': category,  # Используем категорию из JSON
                    'variants': group_data,
                    'image': big_photo,
                    'description': description,
                    'weight': weight,
                })

        wb.close()

    except Exception as e:
        print(f"Ошибка при парсинге Excel: {str(e)}")
        if 'wb' in locals():
            wb.close()

    return products



@app.route('/edit/<int:row>', methods=['GET', 'POST'])
def edit_product(row):
    file_path = os.path.join(app.root_path, app.config['EXCEL_FILE'])
    
    if request.method == 'POST':
        try:
            updated_data = {
                'row': row,
                'Название': request.form['name'],
                'Артикул': request.form['article'],
                'Длина (см)': request.form['length'],
                'Ширина (см)': request.form['width'],
                'Высота (см)': request.form['height'],
                'Цена': request.form['price']
            }
            update_excel_row(file_path, updated_data)
            return redirect(url_for('index'))
        except Exception as e:
            app.logger.error(f"Ошибка при обновлении: {str(e)}")
            return render_template('edit.html', 
                                product=request.form,
                                error=f"Ошибка сохранения: {str(e)}")
    
    try:
        excel_data = get_excel_data(file_path)
        product_to_edit = next((item for item in excel_data if item['row'] == row), None)
        
        if not product_to_edit:
            return "Товар не найден", 404
        
        return render_template('edit.html', product=product_to_edit)
    except Exception as e:
        app.logger.error(f"Ошибка при загрузке: {str(e)}")
        return f"Ошибка: {str(e)}", 500

@app.route('/delete/<int:row>', methods=['POST'])
def delete_product(row):
    file_path = os.path.join(app.root_path, app.config['EXCEL_FILE'])
    delete_excel_row(file_path, row)
    return redirect(url_for('index'))

@app.route('/suppliers')
def suppliers_list():
    if 'user_id' not in session or session.get('user_post', True):
        return redirect(url_for('login'))

    conn = get_db_connection()
    if not conn:
        return "Ошибка подключения к базе данных", 500

    try:
        cursor = conn.cursor(dictionary=True)
        # Получаем поставщиков с информацией о депозите от текущего продавца
        cursor.execute("""
            SELECT 
                u.id, 
                u.username, 
                u.company_info,
                COALESCE((
                    SELECT SUM(amount) 
                    FROM deposits 
                    WHERE seller_id = %s AND supplier_id = u.id
                ), 0) AS deposit_amount
            FROM users u 
            WHERE u.post = 1
        """, (session['user_id'],))
        
        suppliers = []
        for supplier in cursor.fetchall():
            # Парсим company_info если он есть
            if supplier['company_info']:
                try:
                    supplier['company_info'] = json.loads(supplier['company_info'])
                except json.JSONDecodeError:
                    supplier['company_info'] = None
            else:
                supplier['company_info'] = None
            suppliers.append(supplier)
            
        return render_template('suppliers.html', suppliers=suppliers)
    except Error as e:
        print(f"Error fetching suppliers: {e}")
        return "Ошибка сервера", 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

@app.errorhandler(404)
def page_not_found(e):
    return render_template('error.html', message="Страница не найдена"), 404

@app.errorhandler(500)
def internal_server_error(e):
    return render_template('error.html', message="Внутренняя ошибка сервера"), 500

@app.route('/search')
def search_products():
    # Проверка авторизации
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Получаем данные пользователя
    conn = get_db_connection()
    if not conn:
        return render_template('error.html', message="Ошибка подключения к БД"), 500
        
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, username, post FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()
        
        if not user:
            session.pop('user_id', None)
            return redirect(url_for('login'))
        
        search_query = request.args.get('query', '').strip()
        selected_categories = request.args.getlist('category')
        
        # Загрузка данных из Excel
        file_path = os.path.join(app.root_path, app.config['EXCEL_FILE'])
        if not os.path.exists(file_path):
            return render_template('error.html', message="Файл каталога не найден"), 404
        
        products = parse_excel(file_path)
        
        # Фильтрация по поисковому запросу
        if search_query:
            search_lower = search_query.lower()
            products = [p for p in products if search_lower in p['name'].lower()]
        
        # Фильтрация по выбранным категориям
        if selected_categories:
            products = [p for p in products if p.get('category') in selected_categories]
        
        # Получаем список всех уникальных категорий для фильтра
        all_categories = sorted({p.get('category') for p in products if p.get('category')})
        
        # Если пользователь - продавец (post=0) и есть текущий поставщик в сессии
        if not user['post'] and 'current_supplier_id' in session:
            # Получаем данные поставщика
            cursor.execute("SELECT id, username, post FROM users WHERE id = %s AND post = 1", 
                          (session['current_supplier_id'],))
            supplier = cursor.fetchone()
            
            if supplier:
                return render_template('index.html', 
                                    products=products,
                                    user=user,
                                    is_seller=True,
                                    supplier_view=True,
                                    supplier=supplier,
                                    all_categories=all_categories,
                                    selected_categories=selected_categories)
        
        # Для поставщиков (post=1) или если нет текущего поставщика
        return render_template('index.html', 
                            products=products,
                            user=user,
                            is_seller=False,
                            wb_categories=WB_CATEGORIES,
                            supplier_view=False,
                            all_categories=all_categories,
                            selected_categories=selected_categories)
        
    except Exception as e:
        logger.error(f"Ошибка в search_products: {str(e)}")
        return render_template('error.html', message="Внутренняя ошибка сервера"), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

@app.route('/')
def index():
    # Проверка авторизации
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Получаем данные пользователя
    conn = get_db_connection()
    if not conn:
        return render_template('error.html', message="Ошибка подключения к БД"), 500
        
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, username, post FROM users WHERE id = %s", (session['user_id'],))
        user = cursor.fetchone()
        
        if not user:
            session.pop('user_id', None)
            return redirect(url_for('login'))
        
        # Для поставщиков показываем каталог
        if user['post']:
            file_path = os.path.join(app.root_path, app.config['EXCEL_FILE'])
            if not os.path.exists(file_path):
                return render_template('error.html', message="Файл каталога не найден"), 404
            
            try:
                products = parse_excel(file_path)
                # Получаем список всех уникальных категорий для фильтра
                all_categories = sorted({p.get('category') for p in products if p.get('category')})
                
                return render_template('index.html', 
                                    products=products,
                                    user=user,
                                    is_seller=False,
                                    wb_categories=WB_CATEGORIES,
                                    supplier_view=False,
                                    all_categories=all_categories,
                                    selected_categories=[])  # Пустой список, так как фильтры не применены
            except Exception as e:
                logger.error(f"Ошибка загрузки каталога: {str(e)}")
                return render_template('error.html', message="Ошибка загрузки каталога"), 500
        
        # Для продавцов перенаправляем на список поставщиков
        return redirect(url_for('suppliers_list'))
        
    except Exception as e:
        logger.error(f"Ошибка в index: {str(e)}")
        return render_template('error.html', message="Внутренняя ошибка сервера"), 500
    finally:
        if conn and conn.is_connected():
            cursor.close()
            conn.close()

if __name__ == '__main__':
    app.run(debug=True)
